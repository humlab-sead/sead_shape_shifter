import shutil
import tempfile
from pathlib import Path
from typing import Any, Protocol

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from sqlalchemy import create_engine

from src.model import ShapeShiftProject, TableConfig
from src.utility import Registry, create_db_uri, dotget


class DispatchRegistry(Registry[type["Dispatcher"]]):
    """Registry for data store implementations."""

    items: dict[str, type["Dispatcher"]] = {}


Dispatchers: DispatchRegistry = DispatchRegistry()  # pylint: disable=invalid-name


class IDispatcher(Protocol):

    def dispatch(self, target: str, data: dict[str, pd.DataFrame]) -> None: ...


class Dispatcher(IDispatcher):
    """Base class for data dispatchers."""

    def __init__(self, cfg: ShapeShiftProject | dict[str, Any]) -> None:
        self.cfg: ShapeShiftProject = cfg if isinstance(cfg, ShapeShiftProject) else ShapeShiftProject(cfg=cfg)

    @property
    def target_type(self) -> str | None:
        return getattr(self, "_registry_opts", {}).get("target_type", "unknown")

    @property
    def description(self) -> str | None:
        return getattr(self, "_registry_opts", {}).get("description", "")


@Dispatchers.register(key="csv", target_type="folder", description="Dispatch data as CSV files to a folder", extension=None)
class CsvDispatcher(Dispatcher):
    """Dispatcher for CSV data."""

    def dispatch(self, target: str, data: dict[str, pd.DataFrame]) -> None:
        output_dir = Path(target)
        output_dir.mkdir(parents=True, exist_ok=True)
        for entity_name, table in data.items():
            table.to_csv(output_dir / f"{entity_name}.csv", index=False)


@Dispatchers.register(key="zipcsv", target_type="file", description="Dispatch data as CSV files inside a ZIP archive", extension=".zip")
class ZipCsvDispatcher(CsvDispatcher):
    """Dispatcher for CSV data."""

    def dispatch(self, target: str, data: dict[str, pd.DataFrame]) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            super().dispatch(temp_dir, data)
            filename: Path = Path(target)
            filename.parent.mkdir(parents=True, exist_ok=True)
            shutil.make_archive(base_name=str(filename.with_suffix("")), format="zip", root_dir=str(temp_dir))


@Dispatchers.register(key="xlsx", target_type="file", description="Dispatch data as Excel file", extension=".xlsx")
class ExcelDispatcher(Dispatcher):
    """Dispatcher for Excel data."""

    def dispatch(self, target: str, data: dict[str, pd.DataFrame]) -> None:
        with pd.ExcelWriter(target, engine="openpyxl", mode="w") as writer:
            if hasattr(writer, "book"):
                writer.book.calculation.calcMode = "manual"  # type: ignore
            for entity_name in sorted(data):
                data[entity_name].to_excel(writer, sheet_name=entity_name, index=False)


@Dispatchers.register(key="openpyxl", target_type="file", description="Dispatch data as Excel file using openpyxl", extension=".xlsx")
class OpenpyxlExcelDispatcher(Dispatcher):
    """Dispatcher for Excel data using openpyxl."""

    column_colors: dict[str, str] = {
        "header": "#e7e7ef",
        "key_column": "#ccc0da",
        "system_id": "#dce6f1",
        "surrogate_id": "#ebf1de",
        "foreign_key": "#fde9d9",
        "source_column": "#e4dfec",
    }

    def dispatch(self, target: str, data: dict[str, pd.DataFrame]) -> None:
        wb = Workbook(write_only=False)  # Keep False for styling support
        wb.remove(wb.active)  # type: ignore[attr-defined] ; openpyxl creates a default sheet

        # Disable formula calculations for better performance
        wb.calculation.calcMode = "manual"  # type: ignore

        for entity_name in sorted(data):
            table: pd.DataFrame = data[entity_name]
            sheet_name: str = self._safe_sheet_name(entity_name, existing=wb.sheetnames)
            ws = wb.create_sheet(title=sheet_name)

            # Write headers + rows
            for r in dataframe_to_rows(table, index=False, header=True):
                ws.append(r)

            self.style_sheet_columns(entity_name, table, ws)
            self.auto_size_columns(ws)

        wb.save(target)

    @staticmethod
    def _to_argb(color: str) -> str:
        """
        Convert '#RRGGBB' or 'RRGGBB' to openpyxl ARGB 'FFRRGGBB'.
        Accepts 'AARRGGBB' as-is.
        """
        c: str = str(color).strip().lstrip("#")
        if len(c) == 6:  # RRGGBB
            return ("FF" + c).upper()
        if len(c) == 8:  # AARRGGBB
            return c.upper()
        raise ValueError(f"Invalid color: {color!r} (expected RRGGBB or AARRGGBB)")

    @classmethod
    def _solid_fill(cls, color: str) -> PatternFill:
        argb = cls._to_argb(color)
        return PatternFill(start_color=argb, end_color=argb, fill_type="solid")

    def style_sheet_columns(self, entity_name: str, table: pd.DataFrame, ws) -> None:
        # Header row: bright yellow (ARGB)
        header_fill = self._solid_fill(self.column_colors["header"])
        for cell in ws[1]:
            cell.fill = header_fill

        columns: list[str] = table.columns.to_list()
        entity_cfg: TableConfig = self.cfg.get_table(entity_name)

        for column in columns:
            if column in entity_cfg.keys:
                self.set_column_background_color(column, columns, ws, self.column_colors["key_column"])
            elif column == "system_id":
                self.set_column_background_color(column, columns, ws, self.column_colors["system_id"])
            elif column == entity_cfg.surrogate_id:
                self.set_column_background_color(column, columns, ws, self.column_colors["surrogate_id"])
            elif column in entity_cfg.fk_columns:
                self.set_column_background_color(column, columns, ws, self.column_colors["foreign_key"])
            elif column in entity_cfg.safe_columns:
                self.set_column_background_color(column, columns, ws, self.column_colors["source_column"])

    def set_column_background_color(self, column_name: str, columns: list[str], ws, color: str = "D3D3D3") -> None:
        idx: int | None = self.find_column_index(column_name, columns)
        if idx is None:
            return

        fill = self._solid_fill(color)

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=idx, max_col=idx):
            row[0].fill = fill

    def auto_size_columns(self, ws) -> None:
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter  # Get the column name
            for cell in col:
                try:
                    if cell.value is not None:
                        max_length = max(max_length, len(str(cell.value)))
                except Exception:  # pylint: disable=broad-except
                    pass
            ws.column_dimensions[column].width = max_length + 2

    def find_column_index(self, column_name: str, columns: list[str]) -> int | None:
        for idx, col in enumerate(columns):
            if col == column_name:
                return idx + 1  # openpyxl uses 1-based indexing
        return None

    @staticmethod
    def _safe_sheet_name(name: str, existing: list[str]) -> str:
        """Make a string safe for Excel sheet titles and unique within the workbook."""
        clean_name: str = OpenpyxlExcelDispatcher._clean_sheet_name(name)
        if clean_name in existing:
            raise ValueError(f"Duplicate sheet name: {name if name == clean_name else f'{name} -> {clean_name}'}")
        return clean_name

    @staticmethod
    def _clean_sheet_name(name: str) -> str:
        # Excel sheet title rules: max 31 chars, cannot contain : \ / ? * [ ]
        return name.translate(str.maketrans({c: "_" for c in r":\/?*[]"})).strip()[:31] or "Sheet"


@Dispatchers.register(key="db", target_type="database", description="Dispatch data to a database", extension=None)
class DatabaseDispatcher(Dispatcher):
    """Dispatcher for Database data."""

    def dispatch(self, target: str, data: dict[str, pd.DataFrame]) -> None:

        db_opts: dict[str, Any] = dotget(self.cfg.options, "dispatch.database", {}) or {}

        if not db_opts:
            raise ValueError("Database dispatch requires 'dispatch.database' configuration options")

        db_url: str = create_db_uri(**db_opts)

        engine = create_engine(url=db_url)
        with engine.begin() as connection:
            for entity_name, table in data.items():
                table.to_sql(entity_name, con=connection, if_exists="replace", index=False)
