"""Test timezone handling in Excel dispatchers.

This test verifies that timezone-aware datetime columns are properly converted
to timezone-naive datetimes before writing to Excel files, preventing the
"Excel does not support timezones in datetimes" error.
"""

import tempfile
from pathlib import Path

import pandas as pd
import pytest
from openpyxl import load_workbook

from src.dispatch import ExcelDispatcher, OpenpyxlExcelDispatcher

# pylint: disable=redefined-outer-name


@pytest.fixture
def sample_data_with_timezones():
    """Create sample data with timezone-aware datetime columns."""
    return {
        "test_entity": pd.DataFrame(
            {
                "id": [1, 2, 3],
                "name": ["A", "B", "C"],
                "date_updated": pd.to_datetime(
                    ["2024-01-01 10:00:00", "2024-01-02 11:00:00", "2024-01-03 12:00:00"], utc=True
                ),  # UTC timezone
                "date_created": pd.to_datetime(["2023-01-01 10:00:00", "2023-01-02 11:00:00", "2023-01-03 12:00:00"]).tz_localize(
                    "America/New_York"
                ),  # Different timezone (no .dt for DatetimeIndex)
            }
        )
    }


def test_openpyxl_dispatcher_sanitizes_timezones(sample_data_with_timezones):
    """Test that OpenpyxlExcelDispatcher removes timezone info before writing."""
    # Pass minimal project config with test_entity
    cfg = {
        "metadata": {"name": "test", "type": "shapeshifter-project"},
        "entities": {"test_entity": {"type": "fixed", "public_id": "test_id", "keys": [], "columns": ["id", "name"], "values": []}},
    }
    dispatcher = OpenpyxlExcelDispatcher(cfg=cfg)

    with tempfile.TemporaryDirectory() as tmpdir:
        target = str(Path(tmpdir) / "test_output.xlsx")

        # Should not raise "Excel does not support timezones" error
        dispatcher.dispatch(target=target, data=sample_data_with_timezones)

        # Verify file was created and can be read
        assert Path(target).exists()

        # Verify data integrity (datetimes are present, just without timezone)
        wb = load_workbook(target)
        ws = wb["test_entity"]

        # Check headers
        assert ws.cell(1, 1).value == "id"
        assert ws.cell(1, 3).value == "date_updated"

        # Check that datetime values are present (row 2 = first data row)
        date_value = ws.cell(2, 3).value
        assert date_value is not None
        assert isinstance(date_value, pd.Timestamp) or hasattr(date_value, "year")

        wb.close()


def test_excel_dispatcher_sanitizes_timezones(sample_data_with_timezones):
    """Test that ExcelDispatcher removes timezone info before writing."""
    # Pass minimal project config
    cfg = {"metadata": {"name": "test", "type": "shapeshifter-project"}, "entities": {}}
    dispatcher = ExcelDispatcher(cfg=cfg)

    with tempfile.TemporaryDirectory() as tmpdir:
        target = str(Path(tmpdir) / "test_output.xlsx")

        # Should not raise "Excel does not support timezones" error
        dispatcher.dispatch(target=target, data=sample_data_with_timezones)

        # Verify file was created
        assert Path(target).exists()

        # Verify data can be read back
        df = pd.read_excel(target, sheet_name="test_entity")
        assert len(df) == 3
        assert "date_updated" in df.columns
        assert df["date_updated"].notna().all()


def test_sanitize_timezones_method():
    """Test the _sanitize_timezones method directly."""
    # Create DataFrame with mixed timezone-aware and naive datetimes
    df = pd.DataFrame(
        {
            "id": [1, 2],
            "tz_aware": pd.to_datetime(["2024-01-01", "2024-01-02"], utc=True),
            "tz_naive": pd.to_datetime(["2024-01-01", "2024-01-02"]),
            "regular_col": ["A", "B"],
        }
    )

    # Verify input has timezone (tz_aware column)
    assert getattr(df["tz_aware"].dtype, "tz", None) is not None
    # Verify naive column has no timezone (may not have .tz attribute)
    assert getattr(df["tz_naive"].dtype, "tz", None) is None

    # Sanitize
    sanitized = OpenpyxlExcelDispatcher._sanitize_timezones(df)

    # Verify timezone removed from tz_aware column
    assert getattr(sanitized["tz_aware"].dtype, "tz", None) is None
    # Verify naive column unchanged
    assert getattr(sanitized["tz_naive"].dtype, "tz", None) is None
    # Verify regular column unchanged
    assert sanitized["regular_col"].dtype == object

    # Verify original DataFrame unchanged
    assert getattr(df["tz_aware"].dtype, "tz", None) is not None


def test_sanitize_preserves_datetime_values():
    """Test that sanitization preserves the datetime values themselves."""
    df = pd.DataFrame({"date_col": pd.to_datetime(["2024-01-15 14:30:00"], utc=True)})

    original_value = df["date_col"].iloc[0]

    sanitized = OpenpyxlExcelDispatcher._sanitize_timezones(df)
    sanitized_value = sanitized["date_col"].iloc[0]

    # Timezone should be removed but datetime value preserved
    assert hasattr(original_value, "tz") and original_value.tz is not None
    assert not hasattr(sanitized_value, "tz") or sanitized_value.tz is None

    # Year, month, day, hour, minute should be the same
    assert sanitized_value.year == original_value.year
    assert sanitized_value.month == original_value.month
    assert sanitized_value.day == original_value.day
    assert sanitized_value.hour == original_value.hour
    assert sanitized_value.minute == original_value.minute


if __name__ == "__main__":
    pytest.main([__file__, "-v"])
