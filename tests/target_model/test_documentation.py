"""Unit tests for src/target_model/documentation.py."""

from __future__ import annotations

import io
from pathlib import Path

import openpyxl
import pytest

from src.target_model.documentation import (
    DOCUMENT_GENERATORS,
    DocumentFormat,
    ExcelGenerator,
    HTMLDocumentGenerator,
    MarkdownDocumentGenerator,
    SimsDocumentGenerator,
    TargetModelDocumentGenerator,
)
from src.target_model.models import EntitySpec, TargetModel

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def minimal_model(**extra_entities) -> TargetModel:
    """Build a minimal TargetModel with optional entity overrides."""
    return TargetModel.model_validate(
        {
            "model": {"name": "Test Model", "version": "0.1.0"},
            "entities": extra_entities,
            "constraints": [],
        }
    )


def entity_raw(role: str | None = None, **kwargs) -> dict:
    payload: dict = {}
    if role:
        payload["role"] = role
    payload.update(kwargs)
    return payload


# ---------------------------------------------------------------------------
# DocumentFormat
# ---------------------------------------------------------------------------


class TestDocumentFormat:
    def test_all_expected_values_present(self):
        assert set(DocumentFormat) == {
            DocumentFormat.HTML,
            DocumentFormat.MARKDOWN,
            DocumentFormat.EXCEL,
            DocumentFormat.SIMS,
        }

    def test_string_values(self):
        assert DocumentFormat.HTML == "html"
        assert DocumentFormat.MARKDOWN == "markdown"
        assert DocumentFormat.EXCEL == "excel"
        assert DocumentFormat.SIMS == "sims"


# ---------------------------------------------------------------------------
# DOCUMENT_GENERATORS registry
# ---------------------------------------------------------------------------


class TestDocumentGeneratorsRegistry:
    def test_all_formats_registered(self):
        for fmt in DocumentFormat:
            assert DOCUMENT_GENERATORS.get(fmt) is not None


# ---------------------------------------------------------------------------
# ExcelGenerator
# ---------------------------------------------------------------------------


class TestExcelGenerator:
    def _model_with_entities(self) -> TargetModel:
        return minimal_model(
            site=entity_raw(
                role="fact",
                required=True,
                domains=["core"],
                target_table="tbl_sites",
                public_id="site_id",
                columns={"site_name": {"required": True, "type": "string"}, "latitude": {"required": False, "type": "float"}},
                foreign_keys=[{"entity": "location", "required": True}],
            ),
            location=entity_raw(role="lookup", columns={"location_name": {"required": True, "type": "string"}}),
        )

    def test_generate_returns_bytes(self):
        model = self._model_with_entities()
        result = ExcelGenerator().generate(model)
        assert isinstance(result, bytes)
        assert len(result) > 0

    def test_output_is_valid_xlsx(self):

        model = self._model_with_entities()
        result = ExcelGenerator().generate(model)
        wb = openpyxl.load_workbook(io.BytesIO(result))
        assert "Entities" in wb.sheetnames
        assert "Columns" in wb.sheetnames
        assert "Relationships" in wb.sheetnames

    def test_entities_sheet_has_correct_row_count(self):

        model = self._model_with_entities()
        result = ExcelGenerator().generate(model)
        wb = openpyxl.load_workbook(io.BytesIO(result))
        ws = wb["Entities"]
        # 1 header + 2 entity rows
        assert ws.max_row == 3

    def test_used_in_project_column_without_project(self):

        model = self._model_with_entities()
        result = ExcelGenerator().generate(model)
        wb = openpyxl.load_workbook(io.BytesIO(result))
        ws = wb["Entities"]
        # Find "Used in Project" column index (1-based)
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        col_idx = headers.index("Used in Project") + 1
        values = {ws.cell(row=r, column=col_idx).value for r in range(2, ws.max_row + 1)}  # type: ignore[union-attr]
        assert values == {"No"}

    def test_generate_with_empty_model_produces_header_only_sheets(self):

        model = minimal_model()
        result = ExcelGenerator().generate(model)
        wb = openpyxl.load_workbook(io.BytesIO(result))
        assert wb["Entities"].max_row == 1  # header only


# ---------------------------------------------------------------------------
# TextDocumentGenerator._prepare_model_data
# ---------------------------------------------------------------------------


class TestPrepareModelData:
    """Test the shared data preparation used by Markdown and HTML generators."""

    def _generator(self) -> MarkdownDocumentGenerator:
        return MarkdownDocumentGenerator()

    def _model(self) -> TargetModel:
        return minimal_model(
            site=entity_raw(role="fact", required=True, domains=["core", "spatial"]),
            location=entity_raw(role="lookup", required=False, domains=["core"]),
            bridge_entity=entity_raw(role="bridge", required=False, domains=["core"]),
        )

    def test_stats_total_entities(self):
        data = self._generator()._prepare_model_data(self._model())
        assert data["stats"]["total_entities"] == 3

    def test_stats_required_count(self):
        data = self._generator()._prepare_model_data(self._model())
        assert data["stats"]["required_entities"] == 1
        assert data["stats"]["optional_entities"] == 2

    def test_stats_domain_count(self):
        data = self._generator()._prepare_model_data(self._model())
        # "core", "spatial" → 2 distinct domains
        assert data["stats"]["domain_count"] == 2

    def test_by_domain_keys(self):
        data = self._generator()._prepare_model_data(self._model())
        assert "core" in data["by_domain"]
        assert "spatial" in data["by_domain"]

    def test_entities_in_domain_sorted(self):
        data = self._generator()._prepare_model_data(self._model())
        core_names = [e["name"] for e in data["by_domain"]["core"]]
        assert core_names == sorted(core_names)

    def test_used_count_without_project(self):
        data = self._generator()._prepare_model_data(self._model())
        assert data["stats"]["used_count"] == 0
        assert data["has_project_context"] is False
        assert data["project_name"] is None

    def test_total_fks_zero_for_entities_without_fks(self):
        data = self._generator()._prepare_model_data(self._model())
        assert data["stats"]["total_fks"] == 0

    def test_total_columns_zero_for_entities_without_columns(self):
        data = self._generator()._prepare_model_data(self._model())
        assert data["stats"]["total_columns"] == 0

    def test_entity_with_no_domain_goes_to_general(self):
        model = minimal_model(orphan=entity_raw())
        data = self._generator()._prepare_model_data(model)
        assert "general" in data["by_domain"]
        names = [e["name"] for e in data["by_domain"]["general"]]
        assert "orphan" in names


# ---------------------------------------------------------------------------
# MarkdownDocumentGenerator
# ---------------------------------------------------------------------------


class TestMarkdownDocumentGenerator:
    def test_generate_returns_bytes(self):
        model = minimal_model(site=entity_raw(role="fact"))
        result = MarkdownDocumentGenerator().generate(model)
        assert isinstance(result, bytes)
        assert len(result) > 0

    def test_output_is_utf8_decodable(self):
        model = minimal_model(site=entity_raw(role="fact"))
        result = MarkdownDocumentGenerator().generate(model)
        decoded = result.decode("utf-8")
        assert isinstance(decoded, str)


# ---------------------------------------------------------------------------
# HTMLDocumentGenerator
# ---------------------------------------------------------------------------


class TestHTMLDocumentGenerator:
    def test_generate_returns_bytes(self):
        model = minimal_model(site=entity_raw(role="fact"))
        result = HTMLDocumentGenerator().generate(model)
        assert isinstance(result, bytes)
        assert len(result) > 0

    def test_output_contains_html_markers(self):
        model = minimal_model(site=entity_raw(role="fact"))
        decoded = HTMLDocumentGenerator().generate(model).decode("utf-8")
        assert "<" in decoded and ">" in decoded


# ---------------------------------------------------------------------------
# SimsDocumentGenerator._resolve_effective_sims
# ---------------------------------------------------------------------------


class TestResolveEffectiveSims:
    resolve = staticmethod(SimsDocumentGenerator._resolve_effective_sims)

    # --- explicit overrides are preserved ---

    def test_explicit_identity_tracking_is_not_overridden(self):
        spec = EntitySpec(role="fact", identity_tracking="reconciled")
        result = self.resolve(spec)
        assert result["identity_tracking"] == "reconciled"

    def test_explicit_reconciliation_is_not_overridden(self):
        spec = EntitySpec(role="fact", reconciliation="reconcile-fuzzy")
        result = self.resolve(spec)
        assert result["reconciliation"] == "reconcile-fuzzy"

    def test_explicit_aggregate_parent_preserved(self):
        spec = EntitySpec(role="fact", aggregate_parent="site")
        result = self.resolve(spec)
        assert result["aggregate_parent"] == "site"

    # --- aggregate_parent triggers 'child' ---

    def test_aggregate_parent_sets_child_tracking(self):
        spec = EntitySpec(aggregate_parent="sample")
        result = self.resolve(spec)
        assert result["identity_tracking"] == "child"
        assert result["reconciliation"] is None

    # --- role-based defaults ---

    def test_fact_role_defaults_to_tracked_and_allocate(self):
        spec = EntitySpec(role="fact")
        result = self.resolve(spec)
        assert result["identity_tracking"] == "tracked"
        assert result["reconciliation"] == "allocate"

    def test_lookup_role_defaults_to_reconciled_and_reconcile_exact(self):
        spec = EntitySpec(role="lookup")
        result = self.resolve(spec)
        assert result["identity_tracking"] == "reconciled"
        assert result["reconciliation"] == "reconcile-exact"

    def test_classifier_role_defaults_to_reconciled_and_lookup_only(self):
        spec = EntitySpec(role="classifier")
        result = self.resolve(spec)
        assert result["identity_tracking"] == "reconciled"
        assert result["reconciliation"] == "lookup-only"

    def test_bridge_role_defaults_to_derived_and_derive(self):
        spec = EntitySpec(role="bridge")
        result = self.resolve(spec)
        assert result["identity_tracking"] == "derived"
        assert result["reconciliation"] == "derive"

    # --- no role, no overrides ---

    def test_no_role_or_parent_returns_all_none(self):
        spec = EntitySpec()
        result = self.resolve(spec)
        assert result["identity_tracking"] is None
        assert result["reconciliation"] is None
        assert result["aggregate_parent"] is None

    # --- edge: explicit identity_tracking without reconciliation gets default recon ---

    def test_explicit_tracked_without_reconciliation_defaults_to_allocate(self):
        spec = EntitySpec(identity_tracking="tracked")
        result = self.resolve(spec)
        assert result["reconciliation"] == "allocate"

    def test_explicit_derived_without_reconciliation_defaults_to_derive(self):
        spec = EntitySpec(identity_tracking="derived")
        result = self.resolve(spec)
        assert result["reconciliation"] == "derive"

    def test_explicit_child_tracking_sets_reconciliation_none(self):
        spec = EntitySpec(identity_tracking="child")
        result = self.resolve(spec)
        assert result["reconciliation"] is None


# ---------------------------------------------------------------------------
# SimsDocumentGenerator._classify_sims_subtype
# ---------------------------------------------------------------------------


class TestClassifySimsSubtype:
    classify = staticmethod(SimsDocumentGenerator._classify_sims_subtype)

    def test_aggregate_parent_returns_provider_owned_children(self):
        assert (
            self.classify({"identity_tracking": "child", "reconciliation": None, "aggregate_parent": "site"}) == "provider_owned_children"
        )

    def test_tracked_returns_provider_owned_roots(self):
        assert (
            self.classify({"identity_tracking": "tracked", "reconciliation": "allocate", "aggregate_parent": None})
            == "provider_owned_roots"
        )

    def test_reconciled_exact_returns_provider_extensible(self):
        assert (
            self.classify({"identity_tracking": "reconciled", "reconciliation": "reconcile-exact", "aggregate_parent": None})
            == "provider_extensible"
        )

    def test_reconciled_fuzzy_returns_provider_extensible(self):
        assert (
            self.classify({"identity_tracking": "reconciled", "reconciliation": "reconcile-fuzzy", "aggregate_parent": None})
            == "provider_extensible"
        )

    def test_lookup_only_returns_sead_administered(self):
        assert (
            self.classify({"identity_tracking": "reconciled", "reconciliation": "lookup-only", "aggregate_parent": None})
            == "sead_administered"
        )

    def test_lookup_extensible_returns_sead_administered(self):
        assert (
            self.classify({"identity_tracking": "reconciled", "reconciliation": "lookup-extensible", "aggregate_parent": None})
            == "sead_administered"
        )

    def test_derived_returns_bridges(self):
        assert self.classify({"identity_tracking": "derived", "reconciliation": "derive", "aggregate_parent": None}) == "bridges"

    def test_none_identity_tracking_returns_unclassified(self):
        assert self.classify({"identity_tracking": None, "reconciliation": None, "aggregate_parent": None}) == "unclassified"


# ---------------------------------------------------------------------------
# SimsDocumentGenerator._prepare_sims_data
# ---------------------------------------------------------------------------


class TestPrepareSImsData:
    def _generator(self) -> SimsDocumentGenerator:
        return SimsDocumentGenerator()

    def _model(self) -> TargetModel:
        return minimal_model(
            site=entity_raw(role="fact"),
            location=entity_raw(role="lookup"),
            method=entity_raw(role="classifier"),
            sample_site=entity_raw(role="bridge"),
            sample_note=entity_raw(aggregate_parent="site"),
        )

    def test_entities_list_sorted_alphabetically(self):
        data = self._generator()._prepare_sims_data(self._model())
        names = [e["name"] for e in data["entities"]]
        assert names == sorted(names)

    def test_entities_have_required_keys(self):
        data = self._generator()._prepare_sims_data(self._model())
        for entity in data["entities"]:
            for key in ("name", "spec", "identity_tracking", "reconciliation", "aggregate_parent", "sims_subtype"):
                assert key in entity

    def test_stats_counts_by_tracking_type(self):
        data = self._generator()._prepare_sims_data(self._model())
        stats = data["stats"]
        assert stats["total_entities"] == 5
        assert stats["tracked_count"] == 1  # site (fact)
        assert stats["reconciled_count"] == 2  # location + method
        assert stats["derived_count"] == 1  # sample_site (bridge)
        assert stats["child_count"] == 1  # sample_note (aggregate_parent)

    def test_groups_have_all_expected_keys(self):
        data = self._generator()._prepare_sims_data(self._model())
        for key in ["provider_owned_roots", "provider_owned_children", "provider_extensible", "sead_administered", "bridges"]:
            assert key in data["groups"]

    def test_group_order_list(self):
        data = self._generator()._prepare_sims_data(self._model())
        assert data["group_order"] == [
            "provider_owned_roots",
            "provider_owned_children",
            "provider_extensible",
            "sead_administered",
            "bridges",
        ]

    def test_aggregates_populated_for_child_entity(self):
        data = self._generator()._prepare_sims_data(self._model())
        # sample_note has aggregate_parent="site"
        assert "site" in data["aggregates"]
        child_names = [e["name"] for e in data["aggregates"]["site"]]
        assert "sample_note" in child_names

    def test_reconciliation_groups_populated(self):
        data = self._generator()._prepare_sims_data(self._model())
        # fact → allocate; lookup → reconcile-exact; classifier → lookup-only; bridge → derive
        assert "allocate" in data["reconciliation_groups"]
        assert "reconcile-exact" in data["reconciliation_groups"]

    def test_reconciliation_strategies_list_non_empty(self):
        data = self._generator()._prepare_sims_data(self._model())
        assert len(data["reconciliation_strategies"]) > 0
        # Each item is a (strategy_key, description) tuple
        for item in data["reconciliation_strategies"]:
            assert len(item) == 2

    def test_model_metadata_passed_through(self):
        data = self._generator()._prepare_sims_data(self._model())
        assert data["model"].name == "Test Model"


# ---------------------------------------------------------------------------
# SimsDocumentGenerator.generate
# ---------------------------------------------------------------------------


class TestSimsDocumentGeneratorGenerate:
    def test_generate_returns_bytes(self):
        model = minimal_model(site=entity_raw(role="fact"))
        result = SimsDocumentGenerator().generate(model)
        assert isinstance(result, bytes)
        assert len(result) > 0

    def test_output_is_valid_utf8(self):
        model = minimal_model(site=entity_raw(role="fact"))
        result = SimsDocumentGenerator().generate(model)
        decoded = result.decode("utf-8")
        assert isinstance(decoded, str)

    def test_output_contains_entity_name(self):
        model = minimal_model(my_unique_entity=entity_raw(role="fact"))
        decoded = SimsDocumentGenerator().generate(model).decode("utf-8")
        assert "my_unique_entity" in decoded


# ---------------------------------------------------------------------------
# TargetModelDocumentGenerator
# ---------------------------------------------------------------------------


class TestTargetModelDocumentGenerator:
    def _model(self) -> TargetModel:
        return minimal_model(site=entity_raw(role="fact"))

    def test_generate_markdown_returns_bytes(self):
        gen = TargetModelDocumentGenerator(self._model())
        result = gen.generate(DocumentFormat.MARKDOWN)
        assert isinstance(result, bytes)
        assert len(result) > 0

    def test_generate_html_returns_bytes(self):
        gen = TargetModelDocumentGenerator(self._model())
        result = gen.generate(DocumentFormat.HTML)
        assert isinstance(result, bytes)
        assert len(result) > 0

    def test_generate_excel_returns_bytes(self):
        gen = TargetModelDocumentGenerator(self._model())
        result = gen.generate(DocumentFormat.EXCEL)
        assert isinstance(result, bytes)
        assert len(result) > 0

    def test_generate_sims_returns_bytes(self):
        gen = TargetModelDocumentGenerator(self._model())
        result = gen.generate(DocumentFormat.SIMS)
        assert isinstance(result, bytes)
        assert len(result) > 0

    def test_generate_raises_value_error_for_unknown_format(self):
        gen = TargetModelDocumentGenerator(self._model())
        with pytest.raises(ValueError, match="Unsupported format"):
            gen.generate("nonexistent_format")  # type: ignore[arg-type]

    def test_write_to_file_creates_file(self, tmp_path: Path):
        gen = TargetModelDocumentGenerator(self._model())
        out = tmp_path / "output.md"
        gen.write_to_file(DocumentFormat.MARKDOWN, out)
        assert out.exists()
        assert out.stat().st_size > 0

    def test_write_to_file_content_matches_generate(self, tmp_path: Path):
        gen = TargetModelDocumentGenerator(self._model())
        out = tmp_path / "output.md"
        gen.write_to_file(DocumentFormat.MARKDOWN, out)
        assert out.read_bytes() == gen.generate(DocumentFormat.MARKDOWN)
